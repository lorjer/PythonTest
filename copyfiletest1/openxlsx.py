#test xlsx
from openpyxl import load_workbook
wb = load_workbook(filename = 'kpi.xlsx');

#获取sheet 列表
#sheet_names = wb.get_sheet_names()
sheet_names = wb.sheetnames
print("Show the sheet list :",sheet_names)
#print("Show current active sheet index",_active_sheet_index)

#选取所需sheet
#ws = wb.get_sheet_by_name(sheet_names[1])
ws = wb.active(1)

print('Show current sheet title:',ws.title)
#print('Show current sheet row num:',ws.num_rows)
#print('Show current sheet column num:',ws.column)